import os
import glob
import re
import configparser
from datetime import date
import openpyxl as excel
from openpyxl.styles import Font, colors, Alignment, Border, Side


class AutoTL:
    """Vytvoří technický list ze šablony, zapíše data a uloží."""

    def __init__(self, file_name, output, metadata, notes):
        self.name = file_name
        self.output = output
        self.meta = metadata
        self.notes = notes
        self.workbook = excel.Workbook()
        self.excel_output()

    def output_meta(self):
        """Zápis metadat do technického listu."""

        meta_sheet = self.workbook[self.workbook.sheetnames[0]]

        for data in self.meta:
            if data in config[self.output]:
                address = config[self.output][data]
                meta_sheet[address].value = self.meta[data]

    def create_sheet_dodatek(self, dodatek_last_row):
        """Vytvoření dodatkového listu."""

        notes_sheet_dodatek = self.workbook.create_sheet(title='Dodatek TL')

        # proměnné obraničení buněk
        thick = Side(border_style="medium", color="000000")
        thin = Side(border_style="thin", color="000000")

        # nastavení šířky sloupců
        notes_sheet_dodatek.column_dimensions['A'].width = '30'
        notes_sheet_dodatek.column_dimensions['B'].width = '70'

        # nastavení výšky řádků, formátování textu a ohraničení buněk
        for cell in range(1, dodatek_last_row + 1):
            notes_sheet_dodatek.row_dimensions[cell].height = 25
            cell = str(cell)
            notes_sheet_dodatek[f'A{cell}'].alignment = Alignment(horizontal="center", vertical="center")
            notes_sheet_dodatek[f'B{cell}'].alignment = Alignment(horizontal="left", vertical="center")
            notes_sheet_dodatek[f'A{cell}'].border = Border(top=thin, left=thick, right=thick, bottom=thin)
            notes_sheet_dodatek[f'B{cell}'].border = Border(top=thin, right=thick, bottom=thin)

        # záhlaví
        notes_sheet_dodatek.merge_cells('A1:B1')
        notes_sheet_dodatek['A1'].font = Font(name=config['pismo']['font'],
                                              size=int(config['pismo']['velikost']) * 1.5,
                                              bold=True, color=colors.BLACK)
        notes_sheet_dodatek['A1'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
        notes_sheet_dodatek['A1'].value = 'Dodatek TL'

        # ohraničení posledního řádku
        notes_sheet_dodatek[f'A{str(dodatek_last_row)}'].border = Border(left=thick, right=thick, bottom=thick)
        notes_sheet_dodatek[f'B{str(dodatek_last_row)}'].border = Border(right=thick, bottom=thick)
        return notes_sheet_dodatek

    @staticmethod
    def notes_cells(sheet, first_row, last_row, tc_column, notes_column):
        """Generátor adres pro zápis poznámek."""

        for row in range(first_row, last_row + 1):
            tc_cell = sheet[tc_column + str(row)]
            note_cell = sheet[notes_column + str(row)]
            yield tc_cell, note_cell

    def output_notes(self):
        """Zápis poznámek do technického listu."""

        notes_sheet = self.workbook[self.workbook.sheetnames[1]]

        # proměnné fontů pro zápis
        font_red = Font(name=config['pismo']['font'], size=int(config['pismo']['velikost']),
                        bold=True, color=colors.RED)
        font_normal = Font(name=config['pismo']['font'], size=int(config['pismo']['velikost']),
                           bold=False, color=colors.BLACK)
        font_bold = Font(name=config['pismo']['font'], size=int(config['pismo']['velikost']),
                         bold=True, color=colors.BLACK)

        if self.output != 'PRIMA_Vystup':

            # generátor adres
            address = self.notes_cells(sheet=notes_sheet,
                                       first_row=int(config[self.output]['Prvni_radek_poznamek']),
                                       last_row=int(config[self.output]['Posledni_radek_poznamek']),
                                       tc_column=config[self.output]['Sloupec_timecode'],
                                       notes_column=config[self.output]['Sloupec_poznamek'])

            for tc, marker, comment in self.notes:
                try:
                    address_tc, address_comment = next(address)
                except StopIteration:
                    # vytvoření dodatkového listu
                    row_index = self.notes.index([tc, marker, comment])
                    new_last_row = len(self.notes[row_index:]) + 1  # +1 kvůli záhlaví tabulky
                    new_notes_sheet = self.create_sheet_dodatek(new_last_row)

                    # vytvoření nového generátoru pro dodatkový list
                    address = self.notes_cells(sheet=new_notes_sheet,
                                               first_row=2,
                                               last_row=new_last_row,
                                               tc_column='A',
                                               notes_column='B')

                    address_tc, address_comment = next(address)

                # zápis poznámek
                if marker == config['markery']['poznamka']:
                    # zápis komentáře bez timecodu
                    address_comment.font = font_bold
                    address_comment.value = comment
                else:
                    # určení fontu podle barvy markeru
                    if marker == config['markery']['cervena']:
                        address_tc.font = font_red
                        address_comment.font = font_red
                    elif marker == config['markery']['cerna'] or marker == config['markery']['tucne pismo']:
                        address_tc.font = font_bold
                        address_comment.font = font_bold
                    else:
                        address_tc.font = font_normal
                        address_comment.font = font_normal
                    # zápis
                    address_tc.value = tc
                    address_comment.value = comment
        else:
            # proměnné pro zápis
            poznamky = ''
            reklamace = ''

            # příprava markerů pro zápis
            for tc, marker, comment in self.notes:
                if marker == config['markery']['poznamka']:
                    poznamky += comment + '\n'
                else:
                    reklamace += f'{tc} - {comment}' + '\n'

            # nastavení fontu a zápis
            notes_sheet[config[self.output]['Poznamky']].font = font_bold
            notes_sheet[config[self.output]['Poznamky']].value = poznamky

            notes_sheet[config[self.output]['Reklamace']].font = font_red
            notes_sheet[config[self.output]['Reklamace']].value = reklamace

    def excel_output(self):
        """Zápis a uložení technického listu."""

        # ověření platného druhu technického listu
        if self.output in config['zkratky']:
            self.output = config['zkratky'][self.output]
        else:
            print(f'{self.name} ...CHYBA! Neplatný druh TL. - {self.output}')
            return None

        self.workbook = excel.load_workbook(f"Data/Templates/{self.output}{config['template']['pripona vzoru']}")
        self.workbook.template = False

        # zápis dat
        self.output_meta()
        self.output_notes()

        self.workbook.save(f"Output/{self.name}{config['template']['pripona tl']}")
        print(f'{self.name} ...hotovo')


class ParseTxt(AutoTL):
    """Parser markerů exportovaných z Avid Media Composeru do txt souboru."""

    def __init__(self, txt_file):
        self.name = ''
        self.encoding = ''
        self.file_name(txt_file)
        self.meta = dict()
        self.date()
        self.notes = []
        self.parse = []
        self.output = ''
        self.parse_txt(txt_file)
        self.filter_parse()
        self.edit_meta()
        super().__init__(self.name, self.output, self.meta, self.notes)

    def date(self):
        """Přidání data do metadat."""

        datum = date.today()
        self.meta["Datum"] = f'{datum.day}.{datum.month}.{datum.year}'

    def file_name(self, file_path):
        """Uložení jména souboru a kontrola kódování."""

        # oddělení názvu souboru
        file_name = os.path.split(file_path)
        file_name = os.path.splitext(file_name[1])
        file_name = file_name[0]

        # ověření prefixu pro určení kódování souboru
        split_name = file_name.split('_', 1)
        if split_name[0] in config['encoding']:
            self.encoding = config['encoding'][split_name[0]]
            self.name = split_name[1]
        else:
            self.encoding = config['encoding']['default']
            self.name = file_name

    def parse_txt(self, txt_file):
        """Parser textového souboru."""

        with open(txt_file, encoding=self.encoding) as file:
            txt = file.read()
        txt = txt.split("\t1\n")

        txt_pattern = re.compile(
            r"""
            (\d{2}:\d{2}:\d{2}:\d{2})                               # timecode 
            \t.*\t                                                  
            (red|green|blue|cyan|magenta|yellow|black|white)        # barva merkeru
            \t                                                      
            ([\s\S]*)                                               # komentář
            """, re.VERBOSE)

        for line in txt:
            if line.isspace() or not line:
                continue
            line = re.sub("\t1$", "", line)  # ořez nechtěných znaků na posledním řádku
            tc, marker, comment = txt_pattern.search(line).groups()
            self.parse.append([tc, marker, comment])

    def parse_meta(self, comment):
        """Parser markeru s matadaty."""

        metadata = comment.split('\n')

        # první řádek markeru s metadaty musí vždy určovat druh technického listu
        self.output = metadata[0]
        for meta in metadata[1:]:
            # vynechání prázdných řádků a komentářů
            if meta.isspace() or not meta or meta[0] == '#':
                continue

            meta = meta.split(':', 1)  # rozdělění řádku na klíč a hodnotu
            meta = [item.strip(' ') for item in meta]  # ořez bílých znaků

            # zápis do slovníku metadat
            if len(meta) == 2 and meta[0] in config['metadata']:
                self.meta[meta[0]] = meta[1]
            elif len(meta) == 1 and meta[0] in config['metadata']:
                self.meta[meta[0]] = config['metadata'][meta[0]]  # přidání výchozí hodnoty, pokud není zadána

    def filter_parse(self):
        """Filtr markerů na metadata a poznámky."""

        # slovník pro spojování markerů
        operator_index = {}

        operator_pattern = re.compile(r"^{}(\d*)".format(config['znak_rozpětí']['znak']))
        for tc, marker, comment in self.parse:
            # proměnné pro spojení markerů
            operator = None
            get_index = False

            # ověření operátoru pro spojení markerů
            if operator_pattern.search(comment):
                operator = operator_pattern.search(comment).group()
                if operator not in operator_index:
                    # první výskyt operátoru
                    get_index = True
                    comment = re.sub(operator_pattern, "", comment)
                else:
                    # při dalším výskytu se tc připojí k předchozímu a dále se nezpracovává
                    self.notes[operator_index[operator]][0] = f"{self.notes[operator_index[operator]][0]} - {tc}"
                    continue

            # přepis zkratky na požadovaný komentář
            comment = comment.strip(' ')
            if comment in config['zkratky']:
                comment = config['zkratky'][comment]

            # uprava komentáře markeru černé
            if marker == config['markery']['cerna']:
                try:
                    comment = config['zkratky']['cerny_marker'] + ' ' + comment
                except KeyError:
                    pass

            # filtr metadat
            if comment in config['metadata']:
                self.meta[comment] = tc
            elif marker == config['markery']['metadata']:
                self.parse_meta(comment)
            else:
                self.notes.append([tc, marker, comment])

            # uložení indexu markeru pro připojení tc při dalším výskytu operátoru
            if get_index:
                operator_index[operator] = self.notes.index([tc, marker, comment])

    def idec(self):
        """Úprava IDEC pro zápis."""

        if "IDEC" in self.meta:
            idec = re.search(r"^(\d{2})/(\d{3})/(\d{5})/(\d{4})$", self.meta["IDEC"])
            if idec:
                self.meta["IDEC_rok"], self.meta["IDEC_kod"], \
                    self.meta["IDEC_porad"], self.meta["IDEC_ep"] = idec.groups()
            else:
                self.meta["IDEC_porad"] = config["metadata"]["IDEC"]
            self.meta.pop("IDEC")

    def vstup_meta(self):
        """Úprava metadat pro zápis do vstupních technických listů (TV Prima i TV Barrandov)."""

        if "Serie" in self.meta:
            self.meta['Nazev_ORIG'] = f"{self.meta['Nazev_ORIG']} {self.meta['Serie']}"
            self.meta.pop("Serie")

        if "EP" in self.meta:
            if "Nazev_EP_ORIG" in self.meta:
                self.meta['Nazev_EP_ORIG'] = f"EP.{self.meta['EP']} - {self.meta['Nazev_EP_ORIG']}"
            else:
                self.meta['Nazev_EP_ORIG'] = f"EP.{self.meta['EP']}"
            self.meta.pop("EP")

        if "in" in self.meta and "out" in self.meta:
            self.meta["Stopaz_poradu"] = f"{self.meta['in']} - {self.meta['out']}"
            if "end" in self.meta:
                self.meta["Stopaz_celkova"] = f"{self.meta['in']} - {self.meta['end']}"
                if "tl" in self.meta:
                    self.meta["Stopaz_textless"] = f"{self.meta['tl']} - {self.meta['end']}"
                    self.meta.pop("tl")
                self.meta.pop("end")
            else:
                self.meta["Stopaz_celkova"] = self.meta["Stopaz_poradu"]
                self.meta["Stopaz_textless"] = config['metadata']['tl']
            self.meta.pop("in")
            self.meta.pop("out")
        else:
            self.meta["Stopaz_poradu"] = config["metadata"]["out"]
            self.meta["Stopaz_celkova"] = config["metadata"]["out"]

    def tvb_vystup_meta(self):
        """Úprava metadat pro zápis výstupního technického listu pro TV Barrandov."""

        self.idec()

        if "Serie" in self.meta:
            self.meta['Nazev_CZ'] = f"{self.meta['Nazev_CZ']} {self.meta['Serie']}"
            self.meta['Nazev_ORIG'] = f"{self.meta['Nazev_ORIG']} {self.meta['Serie']}"
            self.meta.pop("Serie")

        if "EP" in self.meta:
            if "Nazev_EP_CZ" in self.meta:
                self.meta["Nazev_EP_CZ"] = f"EP.{self.meta['EP']} - {self.meta['Nazev_EP_CZ']}"
            else:
                self.meta["Nazev_EP_CZ"] = f"EP.{self.meta['EP']}"
            self.meta['Nazev_ORIG'] = f"{self.meta['Nazev_ORIG']} EP.{self.meta['EP']}"
            self.meta.pop("EP")

        if "Nazev_EP_ORIG" in self.meta:
            self.meta['Nazev_ORIG'] = f"{self.meta['Nazev_ORIG']} - {self.meta['Nazev_EP_ORIG']}"
            self.meta.pop("Nazev_EP_ORIG")

        if "out" in self.meta:
            self.meta["Stopaz_poradu"] = f"00:02:00:00 - {self.meta['out']}"
            self.meta["Stopaz_celkova"] = f"00:01:45:00 - {self.meta['out']}"
            self.meta.pop("out")
        else:
            self.meta["Stopaz_poradu"] = config["metadata"]["out"]
            self.meta["Stopaz_celkova"] = config["metadata"]["out"]

        if "zt" in self.meta:
            self.meta["Stopaz_bez_ZT"] = f"00:02:00:00 - {self.meta['zt']}"
            self.meta.pop("zt")
        else:
            self.meta["Stopaz_bez_ZT"] = config["metadata"]["zt"]

    def prima_vystup_meta(self):
        """Úprava metadat pro zápis výstupního technického listu pro TV Prima."""

        self.idec()

        if "out" in self.meta:
            self.meta["Stopaz_poradu"] = f"00:02:00:00 - {self.meta['out']}"
            self.meta["Duration"] = self.meta["out"][:3] + str(int(self.meta["out"][3:5]) - 2) + self.meta["out"][5:]
            self.meta.pop("out")
        else:
            self.meta["Stopaz_poradu"] = config["metadata"]["out"]
            self.meta["Duration"] = config["metadata"]["out"]

        if "Kvalita_obrazu" in self.meta:
            self.meta[f"Kvalita_obrazu_{self.meta['Kvalita_obrazu']}"] = "X"
            self.meta.pop("Kvalita_obrazu")

        if "Kvalita_zvuku" in self.meta:
            self.meta[f"Kvalita_zvuku_{self.meta['Kvalita_zvuku']}"] = "X"
            self.meta.pop("Kvalita_zvuku")

    def edit_meta(self):
        """Úprava metadat pro výstup."""

        if self.output == 'TVB_Vstup' or self.output == 'PRIMA_Vstup':
            self.vstup_meta()
        elif self.output == 'TVB_Vystup':
            self.tvb_vystup_meta()
        elif self.output == 'PRIMA_Vystup':
            self.prima_vystup_meta()

        # úprava metadat které mají být zapsány velkými písmeny
        for meta in self.meta:
            if meta in config['caps_meta']:
                self.meta[meta] = self.meta[meta].upper()

        # ověření formátu obrazu, technický list pro 19:9 PB je se liší od standartního
        if self.output == 'TVB_Vystup':
            if '16x9_pillarbox' in self.meta:
                self.output = 'TVB_Vystup_PB'


def main():
    for marker_file in glob.glob('Input_TXT/*.txt'):
        ParseTxt(marker_file)
    input('')


# načtení souboru nastavení
config = configparser.ConfigParser()
try:
    config.read('Data/Config.ini', encoding='UTF-8')
except configparser.MissingSectionHeaderError:
    config.read('Data/Config.ini', encoding='UTF-8-sig')

if __name__ == '__main__':
    main()
